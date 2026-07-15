"""Real, permanent regression coverage for Link to File / Brochure PDF /
Floor Plan / High Res Images across EVERY real source file collected in
this project so far.

Added on request (2026-07) after a repeated pattern: UNION's Link to
File self-reference collision, The Workplace Company's Canva-vs-Website
mixup, Knotel's pitch.com-vs-knotel.com mixup, MetSpace's floor-plan-
vs-photo mixup, and Workplace Plus's multi-building segmentation failure
were each fixed only after a NEW file happened to surface them — never
caught by whichever file prompted the PREVIOUS fix, because nothing ran
the full set together. This suite exercises the REAL app.py route
(Flask's own test_client, not a reimplementation) against every real
source file already collected, checking real, hand-verified values for
these four fields specifically — not just "populated," but matching the
actual real link/content confirmed by hand for each, the same standard
applied throughout this project's own investigation.

Run with: python tests/test_links_and_images.py
Run this ALONGSIDE tests/test_examples.py before ever reporting a fix or
deployment as complete — this is the standing verification process for
this whole class of bug now, not a one-off check. Add a new file's own
real, hand-verified expected values here every time a new real source is
introduced, so coverage actually broadens over time instead of staying
frozen at whichever files prompted this suite's own creation.

Uses live network calls for is_floorplan_image_url's own pixel-content
fetches (extraction.html_images) and any live geocoding — this suite is
deliberately NOT the fast/offline kind tests/test_examples.py mostly is;
it trades speed for testing the REAL, deployed code path end to end.
"""
import io
import sys
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv
load_dotenv(ROOT / ".env")

from openpyxl import load_workbook

import app as app_module


def _process(client, filename):
    """POSTs a real file through the actual /api/process route (Flask's
    own test_client — the real route code, no live HTTP server process
    needed) and returns (file_result_dict, workbook_or_None). Downloads
    and parses the resulting spreadsheet too, so callers can inspect
    real cell/hyperlink values, not just the JSON summary."""
    path = ROOT / filename
    with open(path, "rb") as f:
        resp = client.post(
            "/api/process",
            data={"files": (f, path.name)},
            headers={"X-Access-Token": app_module.ACCESS_TOKEN},
            content_type="multipart/form-data",
        )
    data = resp.get_json()
    fr = data["files"][0]
    if fr["status"] != "ok":
        return fr, None

    dl = client.get(
        f"/api/download/{data['batch_id']}/{fr['output_file']}",
        headers={"X-Access-Token": app_module.ACCESS_TOKEN},
    )
    wb = load_workbook(io.BytesIO(dl.data))
    return fr, wb.active


def _row_dict(ws, header_row=1):
    """Yields one dict per data row: {column_name: (value, hyperlink_target)}."""
    headers = [c.value for c in ws[header_row]]
    for row in ws.iter_rows(min_row=header_row + 1):
        yield {headers[i]: (cell.value, cell.hyperlink.target if cell.hyperlink else None) for i, cell in enumerate(row)}


def _find_row(rows, building_substring, floor_substring=None):
    """First row whose Property Address 1 contains `building_substring`
    (case-insensitive) — and, when given, whose Floor/Unit also contains
    `floor_substring` — or None. Substring search, not exact match,
    since the LLM-fallback sources here re-extract Building text fresh
    each run and could shift punctuation/whitespace slightly; the
    image/link VALUES being checked are tied to the source's own fixed
    structure, not to the LLM's exact wording, so this stays a
    meaningful check even if the address text itself varies a little."""
    needle = building_substring.lower()
    floor_needle = floor_substring.lower() if floor_substring else None
    for row in rows:
        addr = (row.get("Property Address 1", (None, None))[0] or "").lower()
        if needle not in addr:
            continue
        if floor_needle:
            floor = (row.get("Floor/Unit", (None, None))[0] or "").lower()
            if floor_needle not in floor:
                continue
        return row
    return None


def _link(row, field):
    """The real value to compare for a field that may be a plain string
    or a hyperlinked cell (write_xlsx renders a hyperlink with display
    text "Here" — the real URL lives in the hyperlink target, not the
    cell value)."""
    value, target = row.get(field, (None, None))
    return target or value or ""


def check_link_to_file_never_self_references(failures, client, filename):
    """Link to File must never point at the batch's own generated
    output spreadsheet — the exact real collision confirmed (2026-07,
    UNION): when the original upload's own extension matched the output
    extension exactly, both resolved to the same on-disk path, and
    write_xlsx silently overwrote the just-copied real source with the
    generated spreadsheet. Checked generically here (output filename
    never appears in the Link to File URL) rather than re-deriving the
    fix's own internal logic, so this catches ANY future regression of
    this class, not just the exact .xlsx-sourced case that caused it."""
    fr, ws = _process(client, filename)
    if fr["status"] != "ok":
        failures.append(f"{filename}: expected status ok, got {fr['status']} ({fr.get('error')})")
        return
    rows = list(_row_dict(ws))
    if not rows:
        failures.append(f"{filename}: expected at least one row to check Link to File on")
        return
    link_target = _link(rows[0], "Link to File")
    if not link_target:
        failures.append(f"{filename}: expected a non-empty Link to File, got blank")
        return
    if fr["output_file"] in link_target:
        failures.append(
            f"{filename}: Link to File ({link_target!r}) points at the batch's own output file "
            f"({fr['output_file']!r}) instead of the real original source — the exact UNION collision bug"
        )


def check_knotel_links_and_images(failures, client):
    filename = "Fw_ Knotel Availability _ 30_06_2026.eml"
    fr, ws = _process(client, filename)
    if fr["status"] != "ok":
        failures.append(f"{filename}: expected status ok, got {fr['status']}")
        return
    rows = list(_row_dict(ws))

    cases = [
        ("1 Finsbury Market", "3rd Floor", "https://knotel.com/offices/london/1-finsbury-market/u/3rd-floor",
         "https://knotel.directus.app/assets/42fbbce9-3e52-4940-9ac8-779218d7f5e8?download="),
        ("Hallmark", "Hallmark 6th Floor", "https://knotel.com/offices/london/hallmark/u/hallmark-6th-floor",
         "https://knotel.directus.app/assets/0a4300ed-6a23-4936-9604-9a5c4e02a2ed?download="),
        ("2 Leonard Circus", "2nd Floor", "https://knotel.com/offices/london/2-leonard-circus/u/2nd-floor",
         "https://knotel.directus.app/assets/066bdf81-e7e7-450f-aaab-57a4be5f874a?download="),
    ]
    for building, floor, expected_brochure, expected_floorplan in cases:
        row = _find_row(rows, building, floor)
        if row is None:
            failures.append(f"{filename}: expected a row for {building!r} ({floor!r}), found none")
            continue
        if _link(row, "Brochure PDF") != expected_brochure:
            failures.append(f"{filename} {building!r}: expected Brochure PDF {expected_brochure!r}, got {_link(row, 'Brochure PDF')!r}")
        if _link(row, "Floor Plan") != expected_floorplan:
            failures.append(f"{filename} {building!r}: expected Floor Plan {expected_floorplan!r}, got {_link(row, 'Floor Plan')!r}")
        if not _link(row, "High Res Images"):
            failures.append(f"{filename} {building!r}: expected a real High Res Images link (knotel.directus.app), got blank")


def check_metspace_normal_links_and_images(failures, client):
    filename = "Fw_ MetSpace Availability Update.eml"
    fr, ws = _process(client, filename)
    if fr["status"] != "ok":
        failures.append(f"{filename}: expected status ok, got {fr['status']}")
        return
    rows = list(_row_dict(ws))

    row = _find_row(rows, "9-10 Market Place")
    if row is None:
        failures.append(f"{filename}: expected a row for '9-10 Market Place', found none")
        return
    if _link(row, "Brochure PDF") != "https://us.list-manage.com/Q6gnhZazI8v?e=095cb98613&c2id=a206dcecbc185bd9a5d5a46b47a3996f":
        failures.append(f"{filename}: unexpected Brochure PDF for '9-10 Market Place': {_link(row, 'Brochure PDF')!r}")
    if _link(row, "Floor Plan") != "https://mcusercontent.com/53e32083f03d0f8f854aea227/images/5df50ebd-470f-ceba-fa1f-aeb0ac63172f.jpg":
        failures.append(f"{filename}: unexpected Floor Plan for '9-10 Market Place': {_link(row, 'Floor Plan')!r}")
    if _link(row, "High Res Images"):
        # Confirmed real (this template's own docstring, extraction.
        # rules.metspace._attach_floor_plans): no second, genuinely-photo
        # image exists per listing in this source — every mcusercontent.
        # com image is a floor plan, not a photo, for THIS template.
        failures.append(f"{filename}: expected High Res Images blank for MetSpace's usual template, got {_link(row, 'High Res Images')!r}")


def check_metspace_office_of_week_links_and_images(failures, client):
    filename = "Fw_ MetSpace - Office Of The Week!.eml"
    fr, ws = _process(client, filename)
    if fr["status"] != "ok":
        failures.append(f"{filename}: expected status ok, got {fr['status']} ({fr.get('error')})")
        return
    rows = list(_row_dict(ws))
    row = _find_row(rows, "44 Pentonville Road")
    if row is None:
        failures.append(f"{filename}: expected a row for '44 Pentonville Road', found none")
        return

    # Confirmed real (2026-07): both a real interior photo and a floor-
    # plan diagram come from the exact same mcusercontent.com domain with
    # no distinguishing alt text — this is exactly what extraction.
    # html_images.is_floorplan_image_url's pixel-content check exists to
    # tell apart. Visually confirmed (2026-07) which image is which.
    if _link(row, "Floor Plan") != "https://mcusercontent.com/53e32083f03d0f8f854aea227/images/81074122-8bac-1b81-a382-af5f8aaee04c.jpg":
        failures.append(f"{filename}: expected the real floor-plan diagram in Floor Plan, got {_link(row, 'Floor Plan')!r}")
    if _link(row, "High Res Images") != "https://mcusercontent.com/53e32083f03d0f8f854aea227/images/457d7a07-d10f-1ac4-67b3-61c9cf7308ff.jpg":
        failures.append(f"{filename}: expected the real interior photo in High Res Images, got {_link(row, 'High Res Images')!r}")
    # KNOWN GAP (2026-07, not yet fixed): the building's own link text
    # ("44 Pentonville Road") isn't recognized as a brochure link by
    # extraction.html_images.is_brochure_link (which requires the link
    # text to actually mention "brochure"/"particulars"/etc), so
    # Brochure PDF stays blank here even though that link resolves to a
    # real (if JS-viewer) Google Drive page. Pinned as the CURRENT real
    # value, not the ideal one — update this the moment that gap is
    # fixed, don't just delete the check.
    if _link(row, "Brochure PDF"):
        failures.append(
            f"{filename}: expected Brochure PDF still blank (known unfixed gap — see this check's own docstring), "
            f"got {_link(row, 'Brochure PDF')!r} — if this is now populated, UPDATE this expectation, don't just delete it"
        )


def check_gpe_links_and_images(failures, client):
    filename = "Fw_ The latest GPE Fully Managed availability – workspaces you won't want to miss..eml"
    fr, ws = _process(client, filename)
    if fr["status"] != "ok":
        failures.append(f"{filename}: expected status ok, got {fr['status']}")
        return
    rows = list(_row_dict(ws))
    row = _find_row(rows, "Dufour")
    if row is None:
        failures.append(f"{filename}: expected a row for \"16 Dufour's Place\", found none")
        return
    if _link(row, "High Res Images") != "https://assets-gbr.mkt.dynamics.com/545b8b8a-11b8-ed11-9a84-6045bdd093a9/digitalassets/images/9151e227-fa37-f111-88b4-6045bdc203e1?ts=639117650999128730":
        failures.append(f"{filename}: unexpected High Res Images for \"16 Dufour's Place\": {_link(row, 'High Res Images')!r}")
    if _link(row, "Floor Plan") or _link(row, "Brochure PDF"):
        failures.append(
            f"{filename}: expected Floor Plan/Brochure PDF blank for GPE (no such links in this source), "
            f"got Floor Plan={_link(row, 'Floor Plan')!r} Brochure PDF={_link(row, 'Brochure PDF')!r}"
        )


def check_kitts_links_and_images(failures, client):
    filename = "Kitt's Availability (External) - Live Availability.pdf"
    fr, ws = _process(client, filename)
    if fr["status"] != "ok":
        failures.append(f"{filename}: expected status ok, got {fr['status']}")
        return
    rows = list(_row_dict(ws))
    row = _find_row(rows, "28 Bruton Street", "1st")
    if row is None:
        failures.append(f"{filename}: expected a row for '28 Bruton Street' ('1st'), found none")
        return

    # CONFIRMED REAL BUG (2026-07, found while building this suite, NOT
    # YET FIXED): extraction.rules.grid reads Floor Plan/High Res Images
    # straight from the PDF table's own cell TEXT (via pdfplumber), which
    # for a hyperlinked cell is just the display text ("Here") — the
    # real underlying PDF link annotation is never resolved, unlike
    # app.py's own _attach_pdf_images (position-based real-image
    # matching) used for other PDF sources. Pinned as the CURRENT real
    # (broken) value specifically so this suite doesn't silently treat
    # "Here" as correct — update this the moment it's actually fixed.
    if _link(row, "Floor Plan") != "Here" or _link(row, "High Res Images") != "Here":
        failures.append(
            f"{filename}: expected the KNOWN unfixed 'Here' placeholder for Floor Plan/High Res Images "
            f"(grid.py doesn't resolve real PDF hyperlinks yet), got Floor Plan={_link(row, 'Floor Plan')!r} "
            f"High Res Images={_link(row, 'High Res Images')!r} — if this is now a real URL, UPDATE this "
            "expectation to the real value, don't just delete it"
        )


def check_bc_links_and_images(failures, client):
    filename = "BC Current Availability.pdf"
    fr, ws = _process(client, filename)
    if fr["status"] != "ok":
        failures.append(f"{filename}: expected status ok, got {fr['status']}")
        return
    rows = list(_row_dict(ws))
    row = _find_row(rows, "10-12 Alie Street")
    if row is None:
        failures.append(f"{filename}: expected a row for '10-12 Alie Street', found none")
        return
    if _link(row, "Brochure PDF") or _link(row, "Floor Plan") or _link(row, "High Res Images"):
        failures.append(
            f"{filename}: expected all three blank (confirmed real — BC's own PDF table has no image data at "
            f"all), got Brochure PDF={_link(row, 'Brochure PDF')!r} Floor Plan={_link(row, 'Floor Plan')!r} "
            f"High Res Images={_link(row, 'High Res Images')!r}"
        )


def check_breezblok_links_and_images(failures, client):
    filename = "John Stow House.pdf"
    fr, ws = _process(client, filename)
    if fr["status"] != "ok":
        failures.append(f"{filename}: expected status ok, got {fr['status']}")
        return
    rows = list(_row_dict(ws))
    row = _find_row(rows, "John Stow House")
    if row is None:
        failures.append(f"{filename}: expected a row for 'John Stow House', found none")
        return
    if not _link(row, "Floor Plan"):
        failures.append(f"{filename}: expected a real Floor Plan link (the real floor-plan diagram page 6), got blank")
    if not _link(row, "High Res Images"):
        failures.append(f"{filename}: expected a real High Res Images gallery link, got blank")


def check_crown_estate_links_and_images(failures, client):
    filename = "Office Space by The Crown Estate - July 2026.pdf"
    fr, ws = _process(client, filename)
    if fr["status"] != "ok":
        failures.append(f"{filename}: expected status ok, got {fr['status']} ({fr.get('error')})")
        return
    rows = list(_row_dict(ws))
    row = _find_row(rows, "11-12 Pall Mall", "3rd Floor")
    if row is None:
        failures.append(f"{filename}: expected a row for '11-12 Pall Mall' ('3rd Floor'), found none")
        return
    floorplan = _link(row, "Floor Plan")
    if "matterport.com" not in floorplan:
        failures.append(f"{filename}: expected a real Matterport 3D-tour Floor Plan link, got {floorplan!r}")
    if not _link(row, "High Res Images"):
        failures.append(f"{filename}: expected a real High Res Images gallery link, got blank")


def check_union_links_and_images(failures, client):
    filename = "UNION - Availability - June 26 - City 2.xlsx"
    fr, ws = _process(client, filename)
    if fr["status"] != "ok":
        failures.append(f"{filename}: expected status ok, got {fr['status']} ({fr.get('error')})")
        return
    rows = list(_row_dict(ws))
    row = _find_row(rows, "9a Devonshire Square")
    if row is None:
        failures.append(f"{filename}: expected a row for '9a Devonshire Square', found none")
        return
    if _link(row, "Brochure PDF") != "https://app.box.com/s/5ln9uri46xhq586qdoskbc37rhrrftr7":
        failures.append(f"{filename}: unexpected Brochure PDF for '9a Devonshire Square': {_link(row, 'Brochure PDF')!r}")
    if _link(row, "Floor Plan") or _link(row, "High Res Images"):
        # Confirmed real: this exact source .xlsx has zero embedded media
        # images at all (checked directly via its own zip structure).
        failures.append(
            f"{filename}: expected Floor Plan/High Res Images blank (confirmed zero embedded images in this "
            f"exact source file), got Floor Plan={_link(row, 'Floor Plan')!r} High Res Images={_link(row, 'High Res Images')!r}"
        )


def check_workplace_company_links_and_images(failures, client):
    filename = "The Workplace Company Availability.xlsx"
    fr, ws = _process(client, filename)
    if fr["status"] != "ok":
        failures.append(f"{filename}: expected status ok, got {fr['status']} ({fr.get('error')})")
        return
    rows = list(_row_dict(ws))
    row = _find_row(rows, "1 Valentine Place")
    if row is None:
        failures.append(f"{filename}: expected a row for '1 Valentine Place', found none")
        return
    brochure = _link(row, "Brochure PDF")
    if "theworkplacecompany.co.uk" not in brochure:
        failures.append(f"{filename}: expected Brochure PDF to point at theworkplacecompany.co.uk (not Canva), got {brochure!r}")
    if _link(row, "Floor Plan") or _link(row, "High Res Images"):
        # Confirmed real: this source's only embedded image (xl/media/
        # image1.png) is the company's own logo wordmark, not a listing
        # photo — verified by extracting and viewing it directly.
        failures.append(
            f"{filename}: expected Floor Plan/High Res Images blank (this source's only embedded image is the "
            f"company's own logo, not a listing photo), got Floor Plan={_link(row, 'Floor Plan')!r} "
            f"High Res Images={_link(row, 'High Res Images')!r}"
        )


def check_workplace_plus_links_and_images(failures, client):
    filename = "Fw_ Workplace Plus - Availability 14th July.eml"
    fr, ws = _process(client, filename)
    if fr["status"] != "ok":
        failures.append(f"{filename}: expected status ok, got {fr['status']} ({fr.get('error')})")
        return
    rows = list(_row_dict(ws))

    cases = [
        ("77 Gracechurch Street", "6th Floor",
         "https://eot.workplaceplus.co.uk/f/a/jeACq79vkuubiYnOXBZcNg~~/AAAHURA~/dImOlh3zggAULTCU17DEcDKAq7HYESoIGsX7-Q-P3SFjvyUhfAqbztZlhdUxIaBBuW4nyYUXr3XM5IotEA_gmzReraV65G0v0ppG2rkELWYftZoNHbQ4o_p6p6n_BC9I7lluKm5BLezzHtcliosjqLYBkoephVuzYdzZduCkE5DIt64dIZbtMI9CTxZ7HHPY_5xwglw0-LpYbEqx_ivSs1pIncZTLuWhG7sH_CE6EGvFTU79dvdPvemNQOLtxDev",
         "https://gallery.eocampaign1.com/3cfb2852-c496-11f0-bc56-0b4df89ec22e%2F019f5f9e-994e-7620-91ae-d02e77351615.jpg"),
        # Both "150 Waterloo Road" floors must share the SAME segment's
        # photo/brochure — confirmed real (2026-07): this source has one
        # real photo per BUILDING, not per floor, and the fix's own
        # consecutive-building-run grouping (extraction.html_images.
        # _enrich_multi_building) must reflect that, not starve the
        # second floor of a repeat building.
        ("150 Waterloo Road", "2nd Floor",
         "https://eot.workplaceplus.co.uk/f/a/tC9ookNmaWCubhYRhEcy0A~~/AAAHURA~/MQpLKzsSlNE12Mwu0U255ZUCYF6Y89YOyfQ1Jns85pF4y5fqf0QUwd35hZUrvrnL5xl0VQSYLFOrK1Pw0TmopxH3Eq54yGSDRq_azjhszNDBw_NFA7Q4ikURzoyDGpwtqX7Fjx9kz6jZBLy6rrJUZ9AyIAxiW2O9tK0GndHEVqPuqjg9IfOcZEnjtcAC95ewpZCpIBg7I8nzSCRESOi4sqG4un9GlXb5bAb4rA6rPSAXDSfJGnQBc0i74RVw-Ipf",
         "https://gallery.eocampaign1.com/3cfb2852-c496-11f0-bc56-0b4df89ec22e%2F019f5bd5-dbf2-7125-96ce-6043c82f678c.jpg"),
        ("150 Waterloo Road", "4th Floor",
         "https://eot.workplaceplus.co.uk/f/a/tC9ookNmaWCubhYRhEcy0A~~/AAAHURA~/MQpLKzsSlNE12Mwu0U255ZUCYF6Y89YOyfQ1Jns85pF4y5fqf0QUwd35hZUrvrnL5xl0VQSYLFOrK1Pw0TmopxH3Eq54yGSDRq_azjhszNDBw_NFA7Q4ikURzoyDGpwtqX7Fjx9kz6jZBLy6rrJUZ9AyIAxiW2O9tK0GndHEVqPuqjg9IfOcZEnjtcAC95ewpZCpIBg7I8nzSCRESOi4sqG4un9GlXb5bAb4rA6rPSAXDSfJGnQBc0i74RVw-Ipf",
         "https://gallery.eocampaign1.com/3cfb2852-c496-11f0-bc56-0b4df89ec22e%2F019f5bd5-dbf2-7125-96ce-6043c82f678c.jpg"),
        ("8 Durweston Street", "Ground & First Floor",
         "https://eot.workplaceplus.co.uk/f/a/OZchnwvXuLMFHG1H5lIUTA~~/AAAHURA~/MmTDcruEtbhc47XtdfNoBuoxXCk0PDZ8b-SODASOgwbnLVTrd0DNVBBZZ99zTq34zqVpooUAaju7QdyW7fpQcovWWZI-Bxfmd3kT1ETNEsRWG4cqZzHkgG3bcAbIhSkRSCpqdJU5Ml1TJvNAgX-jwxenE_YZwAvxD04XmczoiOUArI7qTU2xY2CIUhiw4RfVBpmGWqpcjeWFxtTIZvKEX4oYSmwbzFnNqtVdS8VeAm2gZqNlStuGzV5aT4ePGxUv",
         "https://gallery.eocampaign1.com/3cfb2852-c496-11f0-bc56-0b4df89ec22e%2F019f5bd2-f8b9-78db-a0aa-1418eedf8d8d.jpg"),
    ]
    for building, floor, expected_brochure, expected_highres in cases:
        row = _find_row(rows, building, floor)
        if row is None:
            failures.append(f"{filename}: expected a row for {building!r} ({floor!r}), found none")
            continue
        if _link(row, "Brochure PDF") != expected_brochure:
            failures.append(f"{filename} {building!r} ({floor!r}): expected Brochure PDF {expected_brochure!r}, got {_link(row, 'Brochure PDF')!r}")
        if _link(row, "High Res Images") != expected_highres:
            failures.append(f"{filename} {building!r} ({floor!r}): expected High Res Images {expected_highres!r}, got {_link(row, 'High Res Images')!r}")


def main():
    app_module.app.config["TESTING"] = True
    client = app_module.app.test_client()

    failures = []

    check_link_to_file_never_self_references(failures, client, "UNION - Availability - June 26 - City 2.xlsx")
    check_link_to_file_never_self_references(failures, client, "The Workplace Company Availability.xlsx")

    check_knotel_links_and_images(failures, client)
    check_metspace_normal_links_and_images(failures, client)
    check_metspace_office_of_week_links_and_images(failures, client)
    check_gpe_links_and_images(failures, client)
    check_kitts_links_and_images(failures, client)
    check_bc_links_and_images(failures, client)
    check_breezblok_links_and_images(failures, client)
    check_crown_estate_links_and_images(failures, client)
    check_union_links_and_images(failures, client)
    check_workplace_company_links_and_images(failures, client)
    check_workplace_plus_links_and_images(failures, client)

    if failures:
        print("\nFAILURES:")
        for f in failures:
            print(" -", f)
        sys.exit(1)
    print("\nAll real source files' Link to File / Brochure PDF / Floor Plan / High Res Images values checked OK.")


if __name__ == "__main__":
    main()
