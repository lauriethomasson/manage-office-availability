"""Writes a single source's extracted records to a formatted .xlsx."""
import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from extraction.schema import COLUMNS

LAT_COL_IDX = COLUMNS.index("Lat") + 1

HEADER_FILL = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFFFF")
CURRENCY_COLS = {"Marketing Price (Based on Min Term) PCM", "Marketing Price (Based on Min Term) PSF"}
NUMBER_COLS = {"Size (sq ft)", "Desks (max)"}
COORDINATE_COLS = {"Lat", "Lng"}
# Only Link to Brochure gets the clickable-hyperlink treatment — it's a
# link we generate ourselves to a file we control and persist. Floor Plan
# and High Res Images hold whatever URL (if any) a source document itself
# provided, which isn't reliable yet, so for now they're left as plain
# text rather than clickable links (there is no "Photo Link" column in
# the current schema — only these two — so nothing to change there).
LINK_COLS = {"Link to Brochure"}
# Free-text columns that can run long enough to overflow into neighboring
# cells — wrapped within their own cell instead, with row height grown to fit.
WRAP_COLS = {"Special Features", "Contacts", "Assigned Agents"}
# Every cell is centered on both axes, including the wrapped long-text
# columns (Special Features/Contacts/Assigned Agents) — those just add
# wrap_text on top of the same centering, so alignment is consistent
# across every column rather than a mix of "center" and "top".
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
WRAP_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
LINE_HEIGHT = 15  # approx. points needed per wrapped line at 11pt Calibri


def write_xlsx(path, records, sheet_title="Listings"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] or "Listings"  # Excel sheet name length limit

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT
    ws.freeze_panes = "A2"

    for row_idx, record in enumerate(records, start=2):
        row = [record.get(c, "") for c in COLUMNS]
        ws.append(row)
        # extraction.pipeline stashes this whenever Lat/Lng/Property
        # Postcode were derived via the web-search fallback (not read from
        # the source) — surfaced here as a cell comment so the specific
        # sources a "(Not in source text)" value was based on are visible
        # directly in the spreadsheet, not just the console log.
        sources = record.get("_geocode_sources")
        if sources:
            comment_text = "Address found via web search, based on:\n" + "\n".join(sources)
            ws.cell(row=row_idx, column=LAT_COL_IDX).comment = Comment(comment_text, "manage-office-availability")

    last_row = ws.max_row
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        max_len = len(col_name)
        for row_idx in range(2, last_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            cell.alignment = CENTER_ALIGNMENT
            if col_name in CURRENCY_COLS and isinstance(val, (int, float)) and val != "":
                cell.number_format = "£#,##0.00" if col_name.endswith("PSF") else "£#,##0"
            elif col_name in NUMBER_COLS and isinstance(val, (int, float)) and val != "":
                cell.number_format = "#,##0"
            elif col_name in COORDINATE_COLS and isinstance(val, (int, float)) and val != "":
                cell.number_format = "0.000000"
            elif col_name in LINK_COLS and isinstance(val, str) and val.startswith("http"):
                # Show a short "Here" label instead of the raw URL — the
                # actual link still goes to the real address via
                # cell.hyperlink, only the displayed text changes.
                actual_url = val
                cell.value = "Here"
                cell.hyperlink = actual_url
                cell.font = Font(color="FF0563C1", underline="single")
                val = cell.value
            elif col_name in WRAP_COLS:
                cell.alignment = WRAP_ALIGNMENT
            max_len = max(max_len, len(str(val)) if val is not None else 0)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)

    # wrap_text alone doesn't make Excel grow the row to fit — that's a
    # rendering computation Excel only does when a human triggers "AutoFit
    # Row Height", not on file load. So estimate wrapped line count from the
    # now-final column widths and set row height explicitly.
    wrap_col_letters = [get_column_letter(i) for i, c in enumerate(COLUMNS, start=1) if c in WRAP_COLS]
    for row_idx in range(2, last_row + 1):
        max_lines = 1
        for letter in wrap_col_letters:
            cell = ws[f"{letter}{row_idx}"]
            text = str(cell.value) if cell.value is not None else ""
            if not text:
                continue
            width = ws.column_dimensions[letter].width or 10
            max_lines = max(max_lines, math.ceil(len(text) / max(width, 1)))
        if max_lines > 1:
            ws.row_dimensions[row_idx].height = max_lines * LINE_HEIGHT

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
